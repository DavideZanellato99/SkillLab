"""File exports: the PDF of a single evaluation, the PDF of a simulation
attempt and the Excel of the evaluations report.

Pure builders: plain data in, file bytes out. No DB and no FastAPI here,
so both are unit-testable and the endpoints stay thin.

Come sono vestiti i due PDF sta in pdf_kit: colori, caratteri e riquadri
sono il design dell'applicazione tradotto su carta, e tenerli separati
lascia qui soltanto cosa raccontano i due referti.
"""

import re
from collections.abc import Sequence
from datetime import datetime
from io import BytesIO

from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from openai_service import EVALUATION_CRITERIA
from pdf_kit import (
    BAD,
    BAD_TINT,
    BODY,
    CARD_PAD_X,
    CARD_PAD_Y,
    FAINT,
    GOOD,
    GOOD_TINT,
    HAIRLINE,
    HEAD_FONT,
    INK,
    MUTED,
    SURFACE,
    VIOLET,
    VIOLET_DEEP,
    VIOLET_EDGE,
    VIOLET_TINT,
    Report,
    Rgb,
    tinted,
)
from schemas import (
    ConversationReviewResponse,
    EvaluationReportRow,
    PreviousAttempt,
    SimulationAnswerResult,
)

# Compact criterion names for the spreadsheet header (the full labels are
# sentences); keys are the ones of openai_service.EVALUATION_CRITERIA and
# the names mirror the dashboard table.
CRITERION_SHORT_LABELS = {
    "rispetto_fasi_chiamata": "Fasi",
    "empatia": "Empatia",
    "sicurezza_competenza": "Sicurezza",
    "appropriatezza_linguaggio": "Linguaggio",
    "identificazione_cliente": "Identificazione",
    "comprensione_casistica": "Casistica",
}

MODE_LABELS = {"voice": "Chiamata", "text": "Chat"}
KIND_LABELS = {
    "open": "Risposta aperta",
    "multiple": "Scelta multipla",
    "ordering": "Ordinamento",
    "matching": "Abbinamento",
}

# Le soglie dei colori del voto, gli stessi della dashboard.
_SCORE_MID = (234, 88, 12)


def _score_rgb(score: float) -> Rgb:
    if score >= 7:
        return GOOD
    if score >= 5:
        return _SCORE_MID
    return BAD


def _score_hex(score: float) -> str:
    red, green, blue = _score_rgb(score)
    return f"{red:02X}{green:02X}{blue:02X}"


def _fmt_score(score: float) -> str:
    return f"{score:.1f}".replace(".", ",")


def _fmt_date(value: datetime) -> str:
    return value.strftime("%d/%m/%Y %H:%M")


def _fmt_time(value: datetime) -> str:
    return value.strftime("%H:%M")


def _spoken(content: str) -> str:
    """Il testo di un messaggio senza la coda che descrive il tono.

    Le battute di una chiamata arrivano da Hume con le emozioni in fondo
    fra graffe, che a schermo diventano la riga "Tono" sotto la bolla. Sulla
    carta quelle parole sono termini inglesi grezzi in mezzo a una
    trascrizione italiana, e la trascrizione serve a rileggere cosa è stato
    detto: restano fuori.
    """
    return re.sub(r"\s*\{[^{}]+\}\s*$", "", content).strip()


# ── Il referto di una valutazione ─────────────────────


def _score_block(
    pdf: Report,
    *,
    label: str,
    score: float,
    color: Rgb,
    badge: str = "",
    badge_color: Rgb | None = None,
    notes: Sequence[str] = (),
    summary: str = "",
) -> None:
    """Il voto in grande, la targhetta del confronto accanto e sotto le righe
    che dicono come e' venuto fuori.

    Lo condividono i due referti: una conversazione valutata e un test
    consegnato arrivano da strade diverse ma finiscono nello stesso numero,
    e sulla carta quel numero deve avere la stessa faccia.
    """
    inner_w = pdf.content_w - 2 * CARD_PAD_X - 1.6
    notes_h = 0.0
    pdf.use(size=8.5)
    for note in notes:
        notes_h += pdf.measure(note, inner_w, 4.4)
    summary_h = 0.0
    if summary:
        pdf.use(size=9.5)
        summary_h = pdf.measure(summary, inner_w, 4.9) + 2.5
    height = 2 * CARD_PAD_Y + 4.8 + 12 + notes_h + summary_h
    with pdf.card(height, accent=color) as (x, y, w):
        pdf.use(font=HEAD_FONT, size=8, color=MUTED)
        pdf.set_char_spacing(0.5)
        pdf.cell(w, 4.8, pdf.safe(label.upper()))
        pdf.set_char_spacing(0)
        if badge:
            accent = badge_color or color
            pdf.pill(badge, right=x + w, y=y + 0.6, fg=accent, bg=tinted(accent))
        pdf.set_xy(x, y + 4.8)
        pdf.use(font=HEAD_FONT, size=26, color=color)
        number = _fmt_score(score)
        pdf.cell(pdf.get_string_width(number) + 1.8, 12, number)
        pdf.use(size=11, color=FAINT)
        pdf.cell(20, 12, "/ 10")
        cursor = y + 16.8
        pdf.use(size=8.5, color=MUTED)
        for note in notes:
            pdf.set_xy(x, cursor)
            pdf.multi_cell(inner_w, 4.4, pdf.safe(note), new_x=XPos.LEFT, new_y=YPos.NEXT)
            cursor = pdf.get_y()
        if summary:
            pdf.set_xy(x, cursor + 2.5)
            pdf.use(size=9.5, color=BODY)
            pdf.multi_cell(inner_w, 4.9, pdf.safe(summary), new_x=XPos.LEFT, new_y=YPos.NEXT)


def _review_card(pdf: Report, review: ConversationReviewResponse) -> None:
    """Il verdetto del docente, in violetto come tutto cio' che a schermo
    viene da una persona e non dalla macchina."""
    inner_w = pdf.content_w - 2 * CARD_PAD_X - 1.6
    reason = (review.override_reason or "").strip()
    note = (review.summary_note or "").strip()
    pdf.use(size=9.5, style="B")
    reason_h = pdf.measure(reason, inner_w, 4.7) if reason else 0.0
    pdf.use(size=9.5)
    note_h = pdf.measure(note, inner_w, 4.7) if note else 0.0
    gap = 2.5 if reason and note else 0.0
    height = 2 * CARD_PAD_Y + 5 + reason_h + gap + note_h
    with pdf.card(height, fill=VIOLET_TINT, border=VIOLET_EDGE, accent=VIOLET) as (x, y, w):
        pdf.use(font=HEAD_FONT, size=8, color=VIOLET_DEEP)
        pdf.set_char_spacing(0.4)
        pdf.cell(w, 5, pdf.safe(f"REVISIONE DEL DOCENTE  ·  {review.reviewer_name}"))
        pdf.set_char_spacing(0)
        cursor = y + 5
        if reason:
            pdf.set_xy(x, cursor)
            pdf.use(size=9.5, style="B", color=VIOLET_DEEP)
            pdf.multi_cell(inner_w, 4.7, pdf.safe(reason), new_x=XPos.LEFT, new_y=YPos.NEXT)
            cursor = pdf.get_y() + gap
        if note:
            pdf.set_xy(x, cursor)
            pdf.use(size=9.5, color=BODY)
            pdf.multi_cell(inner_w, 4.7, pdf.safe(note), new_x=XPos.LEFT, new_y=YPos.NEXT)


def _criterion_card(pdf: Report, criterion: dict, previous: PreviousAttempt | None) -> None:
    """Un criterio: titolo, voto, barra, peso e le parole del formatore."""
    score = float(criterion.get("score", 0) or 0)
    color = _score_rgb(score)
    title = str(criterion.get("label") or criterion.get("key") or "")
    comment = str(criterion.get("comment") or "").strip()
    suggestions = str(criterion.get("suggestions") or "").strip()

    sub = f"Peso {criterion.get('weight', '')}%"
    if previous:
        before = previous.criteria_scores.get(str(criterion.get("key", "")))
        if before is not None:
            delta = round(score - before, 1)
            sign = "+" if delta > 0 else ""
            sub += f"  ·  tentativo precedente {_fmt_score(before)} ({sign}{_fmt_score(delta)})"

    inner_w = pdf.content_w - 2 * CARD_PAD_X
    pdf.use(size=10.5, style="B")
    title_h = pdf.measure(title, inner_w - 24, 5.2)
    pdf.use(size=9.5)
    comment_h = pdf.measure(comment, inner_w, 4.7) + 1.5 if comment else 0.0
    suggestions_h = (
        pdf.note_height(suggestions, inner_w, label="Spunti di miglioramento") + 2.5
        if suggestions
        else 0.0
    )
    height = 2 * CARD_PAD_Y + title_h + 5.4 + 4.6 + comment_h + suggestions_h

    with pdf.card(height) as (x, y, w):
        pdf.use(size=10.5, style="B", color=INK)
        pdf.multi_cell(w - 24, 5.2, pdf.safe(title), new_x=XPos.LEFT, new_y=YPos.NEXT)
        pdf.pill(
            f"{_fmt_score(score)} / 10",
            right=x + w,
            y=y + 0.2,
            fg=color,
            bg=tinted(color),
            size=8.5,
        )
        pdf.bar(x, y + title_h + 2, w, score / 10, color)
        pdf.set_xy(x, y + title_h + 5.4)
        pdf.use(size=8.5, color=MUTED)
        pdf.cell(w, 4.6, pdf.safe(sub))
        cursor = y + title_h + 10
        if comment:
            pdf.set_xy(x, cursor)
            pdf.use(size=9.5, color=BODY)
            pdf.multi_cell(w, 4.7, pdf.safe(comment), new_x=XPos.LEFT, new_y=YPos.NEXT)
            cursor = pdf.get_y() + 1.5
        if suggestions:
            pdf.set_xy(x, cursor + 1)
            pdf.note(
                suggestions,
                x=x,
                width=w,
                fg=VIOLET_DEEP,
                bg=VIOLET_TINT,
                label="Spunti di miglioramento",
            )


def _annotation_card(pdf: Report, annotation: dict) -> None:
    """Una nota appuntata a una battuta, con sopra la battuta stessa.

    Senza la riga su cui sta, un appunto come "qui dovevi chiedere il codice
    cliente" sulla carta non indica niente.
    """
    preview = str(annotation.get("message_preview") or "").strip()
    note = str(annotation.get("note") or "").strip()
    reviewer = str(annotation.get("reviewer_name") or "").strip()
    inner_w = pdf.content_w - 2 * CARD_PAD_X - 1.6
    quote_h = pdf.note_height(preview, inner_w, italic=True, size=8.5) + 2.5 if preview else 0.0
    pdf.use(size=9.5)
    note_h = pdf.measure(note, inner_w, 4.7)
    height = 2 * CARD_PAD_Y + quote_h + note_h + (4.4 if reviewer else 0)
    with pdf.card(height, accent=VIOLET) as (x, y, w):
        cursor = y
        if preview:
            pdf.set_xy(x, cursor)
            pdf.note(f"«{preview}»", x=x, width=w, fg=MUTED, bg=SURFACE, italic=True, size=8.5)
            cursor = pdf.get_y() + 2.5
        pdf.set_xy(x, cursor)
        pdf.use(size=9.5, color=BODY)
        pdf.multi_cell(w, 4.7, pdf.safe(note), new_x=XPos.LEFT, new_y=YPos.NEXT)
        if reviewer:
            pdf.use(size=8, color=FAINT)
            pdf.multi_cell(w, 4.4, pdf.safe(reviewer), new_x=XPos.LEFT, new_y=YPos.NEXT)


def _bubble(pdf: Report, *, speaker: str, when: str, said: str, is_operator: bool) -> None:
    """Una battuta come si vede in chat: l'operatore a destra in violetto,
    l'avatar a sinistra sul grigio chiaro."""
    pad = 3.4
    max_w = pdf.content_w * 0.84
    pdf.use(size=9.5)
    # La bolla è larga quanto la battuta, fino a un massimo: il millimetro in
    # più è il margine con cui `multi_cell` decide di andare a capo, e senza
    # quello una frase corta si spezzerebbe a una parola dalla fine.
    box_w = min(max_w, max(pdf.get_string_width(pdf.safe(said)) + 2 * pad + 2.5, 32))
    text_h = pdf.measure(said, box_w - 2 * pad, 4.7)
    height = text_h + 2 * 3.2
    pdf.keep_together(height + 4.6)

    pdf.use(size=7.8, style="B", color=VIOLET if is_operator else MUTED)
    pdf.set_x(pdf.l_margin)
    pdf.cell(
        pdf.content_w,
        4.4,
        pdf.safe(f"{speaker}{when}"),
        align="R" if is_operator else "L",
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    left = pdf.w - pdf.r_margin - box_w if is_operator else pdf.l_margin
    fill = VIOLET_TINT if is_operator else SURFACE
    border = VIOLET_EDGE if is_operator else HAIRLINE
    bubble = pdf.card(
        height, fill=fill, border=border, width=box_w, left=left, pad_x=pad, pad_y=3.2
    )
    with bubble as (_x, _y, w):
        pdf.use(size=9.5, color=BODY)
        pdf.multi_cell(w, 4.7, pdf.safe(said), new_x=XPos.LEFT, new_y=YPos.NEXT)
    pdf.set_x(pdf.l_margin)


def evaluation_pdf(
    *,
    operator_name: str,
    avatar_name: str,
    conversation_title: str,
    mode: str,
    conversation_at: datetime,
    evaluated_at: datetime,
    overall_score: float,
    summary: str,
    criteria: list[dict],
    previous: PreviousAttempt | None,
    review: ConversationReviewResponse | None = None,
    annotations: list[dict] | None = None,
    messages: list[dict] | None = None,
) -> bytes:
    """One evaluation as an A4 PDF the operator can hand to the trainer.

    `criteria` is the stored result shape (key, label, weight, score,
    comment, suggestions); `previous` adds the per-criterion comparison
    with the previous attempt when there is one.

    `review` is the trainer's own verdict, and when it corrects the score
    the corrected number is the one printed large: this document is what a
    student takes to a dispute, so it has to say what the grade IS, with the
    machine's original next to it rather than in its place. `annotations`
    are the notes pinned to single messages, each a dict with `note`,
    `message_preview` and `reviewer_name`.

    `messages` is the transcript, each a dict with `role`, `content` and
    `created_at`: it closes the document on its own pages, because it is
    what the verdict is about and a correction that cannot be checked
    against what was actually said is a number to be taken on trust.
    """
    override = review.override_score if review else None
    final = override if override is not None else overall_score
    pdf = Report(title="Valutazione della conversazione", subtitle="training con avatar")
    pdf.add_page()

    # Chi ha parlato con chi, su che canale e quando
    pdf.meta(
        [
            ("Operatore", operator_name),
            ("Avatar", avatar_name),
            ("Conversazione", conversation_title),
            ("Canale", MODE_LABELS.get(mode, mode)),
            ("Data", _fmt_date(conversation_at)),
            ("Valutata il", _fmt_date(evaluated_at)),
        ]
    )
    pdf.space(5)

    notes: list[str] = []
    if override is not None:
        notes.append(
            f"Punteggio corretto dal docente, la valutazione automatica assegnava "
            f"{_fmt_score(overall_score)} / 10"
        )
    badge, badge_color = "", None
    if previous:
        delta = round(final - previous.overall_score, 1)
        sign = "+" if delta > 0 else ""
        badge = f"{sign}{_fmt_score(delta)} sul tentativo precedente"
        badge_color = GOOD if delta > 0 else BAD if delta < 0 else MUTED
        notes.append(
            f"Tentativo precedente «{previous.title}» del "
            f"{_fmt_date(previous.conversation_at)}: {_fmt_score(previous.overall_score)} / 10"
        )
    _score_block(
        pdf,
        label="Punteggio complessivo",
        score=final,
        color=_score_rgb(final),
        badge=badge,
        badge_color=badge_color,
        notes=notes,
        summary=summary,
    )
    pdf.space(5)

    # La revisione del docente sta sopra il ragionamento della macchina: chi
    # legge deve incontrare il verdetto umano prima dei sei criteri che
    # quell'umano puo' benissimo aver smentito.
    if review and (review.summary_note or review.override_reason):
        _review_card(pdf, review)
        pdf.space(5)

    pdf.section("Criteri di valutazione")
    for criterion in criteria:
        _criterion_card(pdf, criterion, previous)
        pdf.space(3.5)

    if annotations:
        pdf.space(2)
        pdf.section("Note del docente sulla trascrizione")
        for annotation in annotations:
            _annotation_card(pdf, annotation)
            pdf.space(3.5)

    # La conversazione per intero, su pagine sue: chi legge il referto ha
    # davanti prima il giudizio e poi quello di cui parla, e le due cose non
    # si mescolano perché una si consegna e l'altra si consulta.
    if messages:
        pdf.add_page()
        pdf.section("Trascrizione della conversazione")
        pdf.use(size=8.5, color=MUTED)
        pdf.cell(
            0,
            4.6,
            pdf.safe(f"{operator_name} con {avatar_name}"),
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
        pdf.space(2)
        for message in messages:
            said = _spoken(str(message.get("content") or ""))
            if not said:
                continue
            is_operator = message.get("role") == "user"
            at = message.get("created_at")
            _bubble(
                pdf,
                speaker=operator_name if is_operator else avatar_name,
                when=f"  ·  {_fmt_time(at)}" if isinstance(at, datetime) else "",
                said=said,
                is_operator=is_operator,
            )
            pdf.space(2.5)

    return bytes(pdf.output())


# ── Il referto di un test consegnato ──────────────────


def _attempt_score_rgb(score: float) -> Rgb:
    """The simulator's own thresholds, the school ones its badge uses.

    Not `_score_rgb`: a grade that is amber on screen and orange on paper is
    a grade that has to be read twice.
    """
    if score >= 8:
        return GOOD
    if score >= 6:
        return _SCORE_MID
    return BAD


def _fmt_points(points: float) -> str:
    """Points as the result page writes them: 1, 0,4 — no trailing zero."""
    return f"{points:g}".replace(".", ",")


def _fmt_elapsed(elapsed_ms: int) -> str:
    """How long an answer took, the way the result page says it: "8,2s"."""
    return f"{round(elapsed_ms / 100) / 10:g}".replace(".", ",") + "s"


def _option_rows(answer: SimulationAnswerResult) -> list[tuple[str, Rgb, str]]:
    """Le opzioni come si vedevano, ognuna col suo colore e la sua targhetta.

    Sulla carta non c'e' un verde da leggere al volo, quindi "corretta" e
    "risposta data" stanno scritte accanto alla lettera.
    """
    rows = []
    for index, option in enumerate(answer.options):
        is_right = index == answer.correct_option
        is_given = index == answer.selected_option
        color = GOOD if is_right else BAD if is_given else BODY
        marks = []
        if is_right:
            marks.append("corretta")
        if is_given:
            marks.append("risposta data")
        rows.append((f"{chr(65 + index)}.  {option}", color, ", ".join(marks)))
    return rows


def _ordering_rows(answer: SimulationAnswerResult) -> list[tuple[str, Rgb, str]]:
    """I passi nell'ordine in cui sono stati disposti, col posto che avevano.

    Ogni riga e' verde se quel passo stava dove doveva stare e rossa
    altrimenti, e la targhetta dice qual era il suo posto: senza quel numero
    un elenco tutto rosso non insegna niente, perche' non si vede di quanto
    si era sbagliato.
    """
    correct = answer.correct_steps
    rows = []
    for index, step in enumerate(answer.given_steps):
        right_place = index < len(correct) and _same_item(step, correct[index])
        mark = ""
        if not right_place:
            wanted = next(
                (i for i, s in enumerate(correct) if _same_item(s, step)),
                None,
            )
            mark = f"va al {wanted + 1}" if wanted is not None else ""
        rows.append((f"{index + 1}.  {step}", GOOD if right_place else BAD, mark))
    return rows


def _matching_rows(answer: SimulationAnswerResult) -> list[tuple[str, Rgb, str]]:
    """Gli abbinamenti fatti, ognuno col suo esito.

    Si parte dalle coppie giuste e non da quelle proposte, cosi' le voci
    lasciate scoperte compaiono comunque: una voce senza abbinamento e' una
    coppia sbagliata come le altre, e non vederla nell'elenco farebbe
    sembrare la domanda piu' corta di com'era.
    """
    given = {_item_key(p.left): p.right for p in answer.given_pairs}
    rows = []
    for pair in answer.correct_pairs:
        mine = given.get(_item_key(pair.left), "")
        is_right = _same_item(mine, pair.right)
        text = f"{pair.left}  ->  {mine}" if mine else f"{pair.left}  ->  (nessun abbinamento)"
        rows.append((text, GOOD if is_right else BAD, "" if is_right else f"era: {pair.right}"))
    return rows


def _numbered(steps: list[str]) -> list[str]:
    """I passi con il loro numero davanti, per l'elenco dell'ordine giusto."""
    return [f"{index}.  {step}" for index, step in enumerate(steps, start=1)]


def _item_key(value: str) -> str:
    return " ".join(value.split()).casefold()


def _same_item(a: str, b: str) -> bool:
    return _item_key(a) == _item_key(b)


def _answer_rows(answer: SimulationAnswerResult, kind: str) -> list[tuple[str, Rgb, str]]:
    """Il corpo di una domanda corretta, riga per riga, qualunque sia il tipo.

    Le alternative, i passi disposti o gli abbinamenti fatti finiscono tutti
    nella stessa forma (testo, colore, targhetta), cosi' la scheda che li
    stampa e' una sola: sulla carta le tre cose si leggono allo stesso modo,
    un elenco con accanto detto cosa non andava. Le risposte scritte non
    passano di qui, perche' non sono un elenco.
    """
    if kind == "ordering":
        return _ordering_rows(answer)
    if kind == "matching":
        return _matching_rows(answer)
    return _option_rows(answer)


def _is_blank(answer: SimulationAnswerResult, kind: str) -> bool:
    """Se la domanda e' stata lasciata in bianco, secondo il proprio tipo."""
    if kind == "ordering":
        return not answer.given_steps
    if kind == "matching":
        return not answer.given_pairs
    return answer.selected_option is None


def _question_card(pdf: Report, answer: SimulationAnswerResult, *, kind: str) -> None:
    """Una domanda: cosa chiedeva, cosa e' stato risposto, quanto e' valsa."""
    written = kind == "open"
    rows = [] if written else _answer_rows(answer, kind)
    color = GOOD if answer.is_correct else BAD
    title = f"{answer.position}.  {answer.text}"
    earned = f"{_fmt_points(answer.points)} p."
    if answer.elapsed_ms is not None:
        earned += f"  ·  {_fmt_elapsed(answer.elapsed_ms)}"
    # Quanti elementi erano al posto giusto, dove una risposta puo' essere
    # giusta a meta': senza questo, "0,7 p." su una domanda da sei passi e'
    # un numero che non si sa da dove venga.
    elif answer.item_count:
        earned += f"  ·  {answer.matched_count} su {answer.item_count}"

    inner_w = pdf.content_w - 2 * CARD_PAD_X - 1.6
    pdf.use(size=10.5, style="B")
    title_h = pdf.measure(title, inner_w - 30, 5.2)
    body_h = 0.0
    if written:
        given = (answer.answer_text or "").strip()
        body_h += pdf.note_height(
            given or "Domanda lasciata in bianco.",
            inner_w,
            label="Risposta data",
            italic=not given,
        )
        if answer.expected_answer:
            body_h += 2 + pdf.note_height(answer.expected_answer, inner_w, label="Elementi attesi")
        if answer.feedback:
            pdf.use(size=9.5)
            body_h += 2 + pdf.measure(answer.feedback, inner_w, 4.7)
    else:
        pdf.use(size=9.5)
        for text, _, mark in rows:
            mark_w = _mark_width(pdf, mark)
            body_h += pdf.measure(text, inner_w - mark_w, 4.9) + 1.2
        if _is_blank(answer, kind):
            pdf.use(size=9, style="I")
            body_h += pdf.measure("Domanda lasciata in bianco.", inner_w, 4.6)
        if kind == "ordering" and answer.correct_steps:
            body_h += 2 + pdf.note_height(
                _numbered(answer.correct_steps), inner_w, label="Ordine corretto"
            )
    if answer.explanation:
        body_h += 2.5 + pdf.note_height(answer.explanation, inner_w, label="Spiegazione")
    if answer.sources:
        body_h += 2 + pdf.note_height(
            answer.sources, inner_w, label="Estratti dal documento", italic=True, size=8.5
        )
    height = 2 * CARD_PAD_Y + title_h + 2.5 + body_h

    with pdf.card(height, accent=color) as (x, y, w):
        pdf.use(size=10.5, style="B", color=INK)
        pdf.multi_cell(w - 30, 5.2, pdf.safe(title), new_x=XPos.LEFT, new_y=YPos.NEXT)
        pdf.pill(
            earned,
            right=x + w,
            y=y + 0.2,
            fg=color,
            bg=GOOD_TINT if answer.is_correct else BAD_TINT,
            size=8.5,
        )
        cursor = y + title_h + 2.5
        if written:
            given = (answer.answer_text or "").strip()
            pdf.set_xy(x, cursor)
            pdf.note(
                given or "Domanda lasciata in bianco.",
                x=x,
                width=w,
                fg=BODY if given else MUTED,
                bg=SURFACE,
                label="Risposta data",
                italic=not given,
            )
            cursor = pdf.get_y()
            if answer.expected_answer:
                pdf.set_xy(x, cursor + 2)
                pdf.note(
                    answer.expected_answer,
                    x=x,
                    width=w,
                    fg=GOOD,
                    bg=GOOD_TINT,
                    label="Elementi attesi",
                )
                cursor = pdf.get_y()
            if answer.feedback:
                pdf.set_xy(x, cursor + 2)
                pdf.use(size=9.5, color=BODY)
                pdf.multi_cell(w, 4.7, pdf.safe(answer.feedback), new_x=XPos.LEFT, new_y=YPos.NEXT)
                cursor = pdf.get_y()
        else:
            for text, option_color, mark in rows:
                mark_w = _mark_width(pdf, mark)
                pdf.set_xy(x, cursor)
                pdf.use(size=9.5, style="B" if mark else "", color=option_color)
                pdf.multi_cell(w - mark_w, 4.9, pdf.safe(text), new_x=XPos.LEFT, new_y=YPos.NEXT)
                row_h = pdf.get_y() - cursor
                if mark:
                    pdf.pill(
                        mark,
                        right=x + w,
                        y=cursor + 0.2,
                        fg=option_color,
                        bg=tinted(option_color),
                        size=7.5,
                    )
                cursor += row_h + 1.2
            if _is_blank(answer, kind):
                pdf.set_xy(x, cursor)
                pdf.use(size=9, style="I", color=MUTED)
                pdf.multi_cell(
                    w, 4.6, "Domanda lasciata in bianco.", new_x=XPos.LEFT, new_y=YPos.NEXT
                )
                cursor = pdf.get_y()
            # L'ordine giusto per intero, sotto quello proposto. Le
            # targhette dicono di quanto ogni passo era fuori posto, ma la
            # sequenza corretta va letta di seguito per ricordarsela, ed e'
            # la stessa ragione per cui a schermo sta in un blocco suo.
            if kind == "ordering" and answer.correct_steps:
                pdf.set_xy(x, cursor + 2)
                pdf.note(
                    _numbered(answer.correct_steps),
                    x=x,
                    width=w,
                    fg=GOOD,
                    bg=GOOD_TINT,
                    label="Ordine corretto",
                )
                cursor = pdf.get_y()
        if answer.explanation:
            pdf.set_xy(x, cursor + 2.5)
            pdf.note(
                answer.explanation,
                x=x,
                width=w,
                fg=BODY,
                bg=SURFACE,
                label="Spiegazione",
            )
            cursor = pdf.get_y()
        # A schermo i passaggi stanno dietro a un pannello che si apre, sulla
        # carta stanno semplicemente li', perche' un foglio non si apre. Sotto
        # un'intestazione sola, come nel pannello: una domanda puo' fondarsi
        # su piu' punti del manuale, e sono citazioni della stessa cosa.
        if answer.sources:
            pdf.set_xy(x, cursor + 2)
            pdf.note(
                answer.sources,
                x=x,
                width=w,
                fg=VIOLET_DEEP,
                bg=VIOLET_TINT,
                label="Estratti dal documento",
                italic=True,
                size=8.5,
            )
            cursor = pdf.get_y()


def _mark_width(pdf: Report, mark: str) -> float:
    """Lo spazio che la targhetta di un'opzione toglie al testo dell'opzione."""
    if not mark:
        return 0.0
    pdf.use(size=7.5, style="B")
    return pdf.get_string_width(mark) + 8


def simulation_attempt_pdf(
    *,
    operator_name: str,
    operator_email: str,
    simulation_title: str,
    kind: str,
    submitted_at: datetime,
    correct_count: int,
    question_count: int,
    earned_points: float,
    score: float,
    answers: list[SimulationAnswerResult],
) -> bytes:
    """A submitted test as an A4 PDF, question by question.

    The same page the result screen shows, on paper: the grade, then for
    every question what was answered, what it was worth, why, and the
    passage of the document the question comes from. The kinds of test
    differ only in the body of a question — the options with the right one
    marked, the written answer next to the expected one, the steps in the
    order they were put, the pairs as they were matched — exactly as on
    screen.

    Written in the third person even when it is the student downloading it:
    a sheet that says "you" cannot be handed to anybody else, and being
    handed over is what this document is for.
    """
    written = kind == "open"
    partial = kind in ("ordering", "matching")
    pdf = Report(title="Esito del test", subtitle="simulatore tecnico")
    pdf.add_page()

    pdf.meta(
        [
            ("Operatore", operator_name),
            ("Email", operator_email),
            ("Test", simulation_title),
            ("Tipo", KIND_LABELS.get(kind, kind)),
            ("Consegnato il", _fmt_date(submitted_at)),
        ]
    )
    pdf.space(5)

    # Il voto, e sotto le due conte che lo spiegano: quante risposte sono
    # giuste e' una cosa, quanto valevano un'altra, e un sei con otto
    # risposte esatte non si legge senza entrambe.
    if written:
        reason = (
            "proporzionale alla sua completezza: una risposta parziale vale una parte del punto"
        )
    elif partial:
        reason = "la quota di elementi al posto giusto: quattro su cinque valgono otto decimi"
    else:
        reason = "decrescente con il passare del tempo"
    points = "punto" if earned_points == 1 else "punti"
    _score_block(
        pdf,
        label="Voto",
        score=score,
        color=_attempt_score_rgb(score),
        badge=f"{correct_count} su {question_count} corrette",
        badge_color=_attempt_score_rgb(score),
        notes=[
            f"{_fmt_points(earned_points)} {points} su {question_count}, "
            f"perché il valore di ogni risposta è {reason}"
        ],
    )
    pdf.space(5)

    # Una domanda per foglio, dalla seconda in poi. Un test si rilegge una
    # domanda alla volta, e trovarsi la correzione di quella dopo sotto gli
    # occhi mentre si sta ancora leggendo questa non aiuta; la prima resta
    # sotto il voto perche' li' il foglio e' gia' aperto e sprecarlo a meta'
    # non darebbe niente in cambio. Una domanda che sfora si prende la pagina
    # dopo, ma quella successiva ricomincia comunque da un foglio suo.
    pdf.section("Dettaglio delle domande")
    for index, answer in enumerate(answers):
        if index:
            pdf.add_page()
        _question_card(pdf, answer, kind=kind)

    return bytes(pdf.output())


# ── Excel of the evaluations report ───────────────────


def evaluations_report_xlsx(rows: list[EvaluationReportRow]) -> bytes:
    """The evaluations report as a formatted .xlsx (one row per evaluation).

    Styled header, frozen first row, autofilter, date and score formats and
    threshold colors, so the file is ready to read or slice in Excel with
    no cleanup. Rows come out newest first.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Valutazioni"

    criterion_keys = [key for key, _, _ in EVALUATION_CRITERIA]
    # "Voto" is the grade that counts, "Voto AI" the machine's own: they
    # differ exactly on the rows a trainer corrected, and whoever slices
    # this file has to be able to see which ones those are.
    headers = (
        ["Data", "Conversazione", "Canale", "Operatore", "Email", "Organizzazione", "Avatar"]
        + [CRITERION_SHORT_LABELS.get(key, key) for key in criterion_keys]
        + ["Voto", "Voto AI", "Revisione", "Valutata il"]
    )
    widths = [16, 30, 11, 22, 30, 20, 20] + [13] * len(criterion_keys) + [8, 9, 20, 16]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7C3AED")
    for col, (title, width) in enumerate(zip(headers, widths, strict=True), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.row_dimensions[1].height = 22

    for row_idx, row in enumerate(
        sorted(rows, key=lambda r: r.conversation_at, reverse=True), start=2
    ):
        scores = row.criteria
        operator = f"{row.user_nome} {row.user_cognome}".strip() or row.user_email

        sheet.cell(row=row_idx, column=1, value=row.conversation_at.replace(tzinfo=None))
        sheet.cell(row=row_idx, column=2, value=row.conversation_title)
        sheet.cell(row=row_idx, column=3, value=MODE_LABELS.get(row.mode, row.mode))
        sheet.cell(row=row_idx, column=4, value=operator)
        sheet.cell(row=row_idx, column=5, value=row.user_email)
        sheet.cell(row=row_idx, column=6, value=row.organization_name or "")
        sheet.cell(row=row_idx, column=7, value=row.avatar_name)
        for offset, key in enumerate(criterion_keys):
            score = scores.get(key)
            if score is None:
                continue
            cell = sheet.cell(row=row_idx, column=8 + offset, value=score)
            cell.font = Font(color=_score_hex(score))
        overall = sheet.cell(row=row_idx, column=8 + len(criterion_keys), value=row.overall_score)
        overall.font = Font(bold=True, color=_score_hex(row.overall_score))
        sheet.cell(row=row_idx, column=9 + len(criterion_keys), value=row.ai_overall_score)
        sheet.cell(
            row=row_idx,
            column=10 + len(criterion_keys),
            value=(
                "Punteggio corretto"
                if row.has_override
                else "Note del docente"
                if row.has_review
                else ""
            ),
        )
        sheet.cell(
            row=row_idx,
            column=11 + len(criterion_keys),
            value=row.evaluated_at.replace(tzinfo=None),
        )

    date_format = "dd/mm/yyyy hh:mm"
    last_col = len(headers)
    for row_cells in sheet.iter_rows(min_row=2, max_row=max(2, len(rows) + 1)):
        row_cells[0].number_format = date_format
        row_cells[last_col - 1].number_format = date_format
        # Criteria plus the two score columns (Voto, Voto AI)
        for cell in row_cells[7 : 7 + len(criterion_keys) + 2]:
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{max(2, len(rows) + 1)}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
