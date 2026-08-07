"""Export endpoints: the evaluations report as a formatted .xlsx.

The PDF of a single evaluation is covered in test_chat.py, next to the
rest of the conversation lifecycle it belongs to. Qui restano le prove sul
costruttore puro, che non ha bisogno di passare da un endpoint.
"""

from datetime import UTC, datetime
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from pypdf import PdfReader

from exports import evaluation_pdf, simulation_attempt_pdf
from models import ChatConversation, ConversationEvaluation
from pdf_kit import Report
from schemas import SimulationAnswerResult


def _seed_evaluated_conversation(db_session, user, avatar) -> ChatConversation:
    conversation = ChatConversation(
        user_id=user.id, avatar_id=avatar.id, title="Clienti 1", mode="voice"
    )
    db_session.add(conversation)
    db_session.flush()
    db_session.add(
        ConversationEvaluation(
            conversation_id=conversation.id,
            overall_score=7.5,
            result={
                "summary": "sintesi",
                "criteria": [
                    {
                        "key": "empatia",
                        "label": "Empatia",
                        "weight": 15,
                        "score": 7.5,
                        "comment": "",
                        "suggestions": None,
                    }
                ],
            },
        )
    )
    db_session.flush()
    return conversation


def test_evaluations_report_xlsx(admin_client, db_session, standard_user, make_avatar):
    avatar = make_avatar(category="clienti")
    _seed_evaluated_conversation(db_session, standard_user, avatar)

    response = admin_client.get("/api/admin/evaluations-report/export")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert ".xlsx" in response.headers["content-disposition"]

    sheet = load_workbook(BytesIO(response.content))["Valutazioni"]
    header = [cell.value for cell in sheet[1]]
    assert header[:7] == [
        "Data",
        "Conversazione",
        "Canale",
        "Operatore",
        "Email",
        "Organizzazione",
        "Avatar",
    ]
    assert header[-4:] == ["Voto", "Voto AI", "Revisione", "Valutata il"]

    row = [cell.value for cell in sheet[2]]
    assert row[1] == "Clienti 1"
    assert row[2] == "Chiamata"
    assert row[5] == "Org di test"
    assert row[header.index("Empatia")] == 7.5
    assert row[header.index("Voto")] == 7.5
    # Nessuna revisione: il voto che conta è quello della macchina, e la
    # colonna resta vuota (openpyxl rilegge la cella vuota come None)
    assert row[header.index("Voto AI")] == 7.5
    assert not row[header.index("Revisione")]


def test_a_corrected_score_is_what_the_export_reports(
    admin_client, db_session, standard_user, make_avatar
):
    """Il foglio deve dire il voto che lo studente ha ricevuto, non quello
    che la macchina aveva proposto."""
    avatar = make_avatar(category="clienti")
    conversation = _seed_evaluated_conversation(db_session, standard_user, avatar)
    admin_client.put(
        f"/api/admin/conversations/{conversation.id}/review",
        json={"override_score": 9, "override_reason": "Gestione del reclamo impeccabile."},
    )

    response = admin_client.get("/api/admin/evaluations-report/export")

    sheet = load_workbook(BytesIO(response.content))["Valutazioni"]
    header = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    assert row[header.index("Voto")] == 9
    assert row[header.index("Voto AI")] == 7.5
    assert row[header.index("Revisione")] == "Punteggio corretto"


def test_export_is_admin_only(user_client):
    assert user_client.get("/api/admin/evaluations-report/export").status_code == 403


# ── Il PDF, dove i caratteri incorporati fanno la differenza ──────────


def test_il_referto_regge_i_caratteri_di_un_testo_italiano():
    """Le virgolette caporali e gli accenti arrivano sulla pagina come sono.

    Il testo di una valutazione lo scrive un LLM: prima dei caratteri
    incorporati finiva schiacciato sul latin-1, e « » ed em dash diventavano
    punti interrogativi.
    """
    blob = evaluation_pdf(
        operator_name="Mario Rossi",
        avatar_name="Anna",
        conversation_title="Reclamo «carta bloccata» — primo tentativo",
        mode="voice",
        conversation_at=datetime.now(UTC),
        evaluated_at=datetime.now(UTC),
        overall_score=7.5,
        summary="L'operatore ha detto «capisco» al momento giusto.",
        criteria=[
            {
                "key": "empatia",
                "label": "Empatia",
                "weight": 15,
                "score": 7.5,
                "comment": "Tono adeguato — nessuna forzatura.",
                "suggestions": "Riconosci l'emozione prima della procedura.",
            }
        ],
        previous=None,
        messages=[{"role": "user", "content": "Buongiorno, sono Mario 😀", "created_at": None}],
    )

    assert blob.startswith(b"%PDF")


def test_una_risposta_lunghissima_non_sfonda_la_pagina():
    """Una risposta scritta può arrivare a cinquemila caratteri.

    Più di quanto stia in una pagina: la scheda rinuncia alla cornice e il
    testo continua sulla pagina dopo, invece di finire oltre il piede.
    """
    lunga = "Verifico l'identità del cliente e apro la contestazione a sistema. " * 75

    blob = simulation_attempt_pdf(
        operator_name="Mario Rossi",
        operator_email="mario@esempio.it",
        simulation_title="Procedura di rimborso",
        kind="open",
        submitted_at=datetime.now(UTC),
        correct_count=1,
        question_count=1,
        earned_points=1.0,
        score=6.0,
        answers=[
            SimulationAnswerResult(
                question_id=uuid4(),
                position=1,
                text="Descrivi la procedura di rimborso.",
                answer_text=lunga[:5000],
                expected_answer="Verifica, contestazione, blocco, comunicazione dei tempi.",
                feedback="Completa ma ripetitiva.",
                is_correct=True,
                points=1,
                elapsed_ms=180000,
                explanation="",
            )
        ],
    )

    assert blob.startswith(b"%PDF")
    assert blob.count(b"/Type /Page\n") > 1


def test_i_passaggi_del_documento_stanno_sotto_una_intestazione_sola():
    """Una domanda può fondarsi su più punti del manuale.

    Sono citazioni della stessa cosa e stanno sotto un titolo solo, come nel
    pannello a schermo, invece di ripetere l'intestazione sopra ognuna.
    """
    passaggi = [
        "L'identificazione richiede due elementi anagrafici.",
        "Per le operazioni dispositive se ne aggiunge un terzo.",
        "Lo sblocco della carta è un'operazione dispositiva.",
    ]

    blob = simulation_attempt_pdf(
        operator_name="Mario Rossi",
        operator_email="mario@esempio.it",
        simulation_title="Sblocco carta",
        kind="multiple",
        submitted_at=datetime.now(UTC),
        correct_count=1,
        question_count=1,
        earned_points=1.0,
        score=10.0,
        answers=[
            SimulationAnswerResult(
                question_id=uuid4(),
                position=1,
                text="Quale verifica va fatta prima di sbloccare una carta?",
                options=["Due elementi", "Tre elementi"],
                correct_option=1,
                selected_option=1,
                is_correct=True,
                points=1,
                elapsed_ms=8000,
                explanation="Servono tre elementi.",
                sources=passaggi,
            )
        ],
    )

    text = "\n".join(page.extract_text() for page in PdfReader(BytesIO(blob)).pages)
    assert text.count("ESTRATTI DAL DOCUMENTO") == 1
    for passaggio in passaggi:
        assert passaggio.split(".")[0] in text.replace("\n", " ")


def test_i_caratteri_senza_glifo_restano_fuori_dalla_pagina():
    """Un'emoji nel commento non deve diventare un quadratino vuoto."""
    report = Report(title="Prova", subtitle="prova")

    assert report.safe("Bravo 😀 così «bene»") == "Bravo  così «bene»"
